import collections

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Subquery, OuterRef, Sum
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from administracion.models import Profile, Docente, Nivel, GrupoAcademico, Matricula, Periodo, NotaParcial, Calificacion, \
    DocentesGrupoAcademico, TipoDocente, FallaAsistencia, usuarioTieneGrupo, Observacion, getEstadoMatricula, \
    PreinscripcionHorarioCurso, OfertaAcademica
import json, os
from datetime import datetime, date
from administracion.util import CSVWriter, humanizar_fecha
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import load_workbook
from io import BytesIO
import uuid
import tempfile
from PyPDF2 import PdfMerger
from django.urls import reverse
from django.utils.safestring import mark_safe
from xlsx2html import xlsx2html


def isDocente(user):
    persona = Profile.objects.filter(usuario__id = user.id).first()
    return Docente.objects.filter(persona__id = persona.id).exists()


@login_required
def cursosAsociadosList(request):

    user = request.user
    es_administrador = usuarioTieneGrupo(user, 'Administrador') or usuarioTieneGrupo(user, 'Coordinador')

    periodo_actual_id = request.session["periodo_contextualizado_id"]
    try:
        periodo = Periodo.objects.get(pk=periodo_actual_id)
    except Periodo.DoesNotExist:
        periodo = None
    context = {}
    mis_cursos = collections.OrderedDict()

    if request.method == 'GET':
        persona = Profile.objects.filter(usuario__id=user.id).first()
        docentes = Docente.objects.filter(persona__id=persona.id)

        if len(docentes) > 0 or es_administrador:

            if es_administrador:
                grupos = GrupoAcademico.objects.filter(horarioCurso__curso__oferta_academica__periodo=periodo).order_by('horarioCurso__curso__nivel__idioma__nombre', 'horarioCurso__curso__nivel__nombre')
            else:
                grupos = GrupoAcademico.objects.filter(docentesgrupoacademico__docente__in=docentes, horarioCurso__curso__oferta_academica__periodo=periodo).order_by('horarioCurso__curso__nivel__idioma__nombre', 'horarioCurso__curso__nivel__nombre')
            for grupo in grupos:
                curso = grupo.horarioCurso.curso
                # Matriculas: Se excluyen aquellas que estan canceladas o aplazadas por usuario y departamento
                matriculas = Matricula.objects.filter(grupo=grupo).exclude(estado_matricula__in=[4,5,6]).all()
                if curso not in mis_cursos:
                    mis_cursos[curso] = []
                mis_cursos[curso].append([grupo, len(matriculas)])

            if periodo.fecha_calificacion == date.today():
                calificacion_activa = True
            else:
                calificacion_activa = False

            context = {'mis_cursos': mis_cursos, 'periodo': periodo, 'calificacion_activa': calificacion_activa}

        else:
            messages.warning(request, 'Lo sentimos, a usted no le ha sido asignado el rol de docente.'
                             'Por favor comuníquese con el administrador del sistema')

    return render(request, "administracion/docente/mis_cursos_list.html", context)


@login_required
def listadoEstudiantesPorGrupo(request, grupoacademico):

    context = {}
    user = request.user
    persona = Profile.objects.filter(usuario__id=user.id).first()
    docentes = Docente.objects.filter(persona__id=persona.id)
    tipo_general = TipoDocente.objects.get(tipo='General')
    tipo_especializado = TipoDocente.objects.get(tipo='Especializado')
    periodo_actual = request.session["periodo_contextualizado_id"]
    tipo_docente = None
    es_administrador = False
    observaciones = {}

    if request.method == 'GET':
        try:
            grupo = GrupoAcademico.objects.get(pk=grupoacademico)
        except GrupoAcademico.DoesNotExist:
            grupo = None

        try:
            periodo = Periodo.objects.get(pk=periodo_actual)
        except Periodo.DoesNotExist:
            periodo = None

        if grupo and periodo:
            #Matriculas: Se excluyen aquellas que estan canceladas o aplazadas por usuario y departamento
            matriculas = Matricula.objects.filter(grupo=grupo).exclude(estado_matricula__in=[4,5,6]).order_by('estudiante__primer_apellido', 'estudiante__segundo_apellido')
            escala_notas = grupo.horarioCurso.curso.oferta_academica.programa.escala_nota
            docentes_generales = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, tipo_docente=tipo_general).all()
            docentes_especializados = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, tipo_docente=tipo_especializado).all()

            docente = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, docente__in=docentes).first()
            salones = grupo.salones.all()

            if docente:
                tipo_docente = docente.tipo_docente
            if usuarioTieneGrupo(user, 'Administrador'):
                es_administrador = True

            for matricula in matriculas:
                observaciones[matricula] = len(Observacion.objects.filter(matricula=matricula))

            if periodo.fecha_calificacion == date.today():
                calificacion_activa = True
            else:
                calificacion_activa = False

            context = {'matriculas': matriculas, 'periodo': periodo, 'grupo': grupo,
                       'escala_notas': escala_notas, 'calificacion_activa': calificacion_activa,
                       'docentes_generales': docentes_generales, 'docentes_especializados': docentes_especializados,
                       'tipo_docente': tipo_docente, 'es_administrador': es_administrador,
                       'salones': salones, 'observaciones_matricula': observaciones}

    return render(request, "administracion/docente/mis_estudiantes_list.html", context)


@login_required
def listadoCalificacionesPorGrupo(request, grupoacademico):

    context = {}
    user = request.user
    persona = Profile.objects.filter(usuario__id=user.id).first()
    docentes = Docente.objects.filter(persona__id=persona.id)
    tipo_general = TipoDocente.objects.get(tipo='General')
    tipo_especializado = TipoDocente.objects.get(tipo='Especializado')
    periodo_actual = request.session["periodo_contextualizado_id"]
    tipos_docente = None
    es_administrador_o_coordinador = False

    if request.method == 'GET':
        try:
            grupo = GrupoAcademico.objects.get(pk=grupoacademico)
        except GrupoAcademico.DoesNotExist:
            grupo = None

        try:
            periodo = Periodo.objects.get(pk=periodo_actual)
        except Periodo.DoesNotExist:
            periodo = None

        if grupo and periodo:
            calificaciones_matricula = {}
            # Matriculas: Se excluyen aquellas que estan canceladas o aplazadas por usuario y departamento
            preinscripcion = PreinscripcionHorarioCurso.objects.filter(estado_preinscripcion__in=[1, 3, 5],
                                                                       horario_cupo__curso__oferta_academica__periodo=OuterRef(
                                                                           'grupo__horarioCurso__curso__oferta_academica__periodo'),
                                                                       horario_cupo__curso__oferta_academica__programa=OuterRef(
                                                                           'grupo__horarioCurso__curso__oferta_academica__programa'),
                                                                       horario_cupo__curso__nivel=OuterRef(
                                                                           'grupo__horarioCurso__curso__nivel'),
                                                                       persona=OuterRef('estudiante')
                                                                       )

            matriculas = Matricula.objects.filter(
                grupo=grupo
            ).exclude(
                estado_matricula__in=[4, 5, 6]
            ).order_by(
                'estudiante__primer_apellido'
            ).annotate(
                estado_preinscripcion=Subquery(preinscripcion.values('estado_preinscripcion'))
            )
            for matricula in matriculas:
                calificaciones = Calificacion.objects.filter(matricula=matricula).all()
                if matricula not in calificaciones_matricula:
                    calificaciones_matricula[matricula] = {}
                for calificacion in calificaciones:
                    nota_parcial = calificacion.nota
                    calificaciones_matricula[matricula][nota_parcial] = calificacion.calificacion

            conjunto_notas = grupo.horarioCurso.curso.conjunto_notas
            notas = NotaParcial.objects.filter(conjunto_notas=conjunto_notas).order_by('orden_nota_conjunto').all()
            escala_notas = grupo.horarioCurso.curso.oferta_academica.programa.escala_nota
            docentes_generales = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, tipo_docente=tipo_general).all()
            docentes_especializados = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, tipo_docente=tipo_especializado).all()

            docentes = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo, docente__in=docentes).all()

            if docentes:
                tipos_docente = [docente.tipo_docente.id for docente in docentes]
                #tipo_docente = docente.tipo_docente
            if usuarioTieneGrupo(user, 'Administrador') or usuarioTieneGrupo(user, 'Coordinador'):
                es_administrador_o_coordinador = True

            context = {'matriculas': matriculas, 'periodo': periodo, 'grupo': grupo, 'notas': notas, 'range': range(len(notas)),
                       'calificaciones_matricula': calificaciones_matricula, 'escala_notas': escala_notas,
                       'docentes_generales': docentes_generales, 'docentes_especializados': docentes_especializados,
                       'tipos_docente': tipos_docente, 'es_administrador': es_administrador_o_coordinador,
                       'tipo_general': tipo_general}

    return render(request, "administracion/docente/matriculas_list.html", context)


def asignarCalificacion(matricula_nota, nota_parcial, calificacion, escala_notas):

    errores = ''
    try:
        calificacion_nota = float(calificacion)
    except:
        calificacion_nota = 0.0

    try:
        calificacion_encontrada = Calificacion.objects.get(matricula=matricula_nota, nota=nota_parcial)
    except Calificacion.DoesNotExist:
        calificacion_encontrada = None

    if calificacion_nota > escala_notas.nota_maxima:
        return 'La calificación debe ser menor o igual a ' + str(escala_notas.nota_maxima) + ' para el estudiante: ' + str(matricula_nota.estudiante.numero_documento)
    elif calificacion_nota < escala_notas.nota_minima:
        return 'La calificación debe ser mayor igual a ' + str(escala_notas.nota_minima) + ' para el estudiante: ' + str(matricula_nota.estudiante.numero_documento)

    if calificacion_encontrada:

        calificacion_encontrada.calificacion = calificacion_nota
        calificacion_encontrada.save()
    else:
        calificacion = Calificacion(matricula=matricula_nota, nota=nota_parcial, calificacion=calificacion_nota)
        calificacion.save()

    return errores

def asignarCalificacionFinal(matricula_nota, escala_notas):

    calificaciones_matricula = Calificacion.objects.filter(matricula=matricula_nota).all()
    calificacion_final = 0.0
    for calificacion in calificaciones_matricula:
        calificacion_final += calificacion.calificacion * calificacion.nota.ponderacion / 100
    matricula_nota.calificacionFinal = round(calificacion_final, escala_notas.numero_decimales)
    matricula_nota.save()

@login_required
@csrf_exempt
def calificarGrupo(request, grupoacademico):

    errores = ''
    if request.method == 'POST':
        estudiantes = request.POST.get('json_data')
        estudiantes = json.loads(estudiantes)
        count_calificaciones = 0

        try:
            grupo = GrupoAcademico.objects.get(pk=grupoacademico)
        except:
            grupo = None

        if grupo:

            escala_notas = grupo.horarioCurso.curso.oferta_academica.programa.escala_nota

            for estudiante in estudiantes:
                for matricula in estudiante:

                    try:
                        matricula_nota = Matricula.objects.get(pk=matricula)
                    except:
                        matricula_nota = None

                    for nota in estudiante[matricula]:
                        calificacion = estudiante[matricula][nota]

                        try:
                            nota_parcial = NotaParcial.objects.get(pk=nota)
                        except:
                            nota_parcial = None

                        if matricula_nota and nota_parcial:

                            errores += asignarCalificacion(matricula_nota, nota_parcial, calificacion, escala_notas)

                    count_calificaciones += 1
                    asignarCalificacionFinal(matricula_nota, escala_notas)

        status = 1
        message = 'Se han guardado las calificaciones de ' + str(len(estudiantes)) + ' estudiantes del grupo: ' + grupo.nombre

        if errores != '':
            status = 0
            message = (errores)

        response = {'status': status, 'message': message}
        return HttpResponse(json.dumps(response), content_type='application/json')

    return render(request, 'administracion/docente/mis_cursos_list.html')

def actualizarEstadoMatriculasPorGrupo(grupo_academico, ofertas_periodo):

    matriculas = Matricula.objects.filter(grupo=grupo_academico).all()
    curso = grupo_academico.horarioCurso.curso
    numero_maximo_fallas = curso.nivel.fallas_maximas
    nota_aprobatoria = curso.oferta_academica.programa.escala_nota.nota_aprobatoria

    for matricula in matriculas:
        fallas = FallaAsistencia.objects.filter(matricula=matricula).aggregate(Sum('cantidad_fallas'))['cantidad_fallas__sum']

        try:
            preinscripcion_asociada = PreinscripcionHorarioCurso.objects.get(persona=matricula.estudiante,
                                                                         horario_cupo__curso=matricula.grupo.horarioCurso.curso,
                                                                         estado_preinscripcion__in=[1,3],
                                                                         horario_cupo__curso__oferta_academica__in=ofertas_periodo)  # Inscrito o Pendiente
        except PreinscripcionHorarioCurso.DoesNotExist:
            preinscripcion_asociada = None

        if fallas:
            matricula.total_fallas = fallas

        if matricula.estado_matricula not in [4, 5, 6]:  # Cancelado, Aplazado por usuario, Aplazado por departamento
            if matricula.total_fallas > numero_maximo_fallas:
                matricula.estado_matricula = 8  # Reprobado por inasistencia
            else:
                matricula.estado_matricula = 7  # en curso

        if matricula.estado_matricula not in [8, 4, 5, 6]:
            if matricula.calificacionFinal < nota_aprobatoria:
                matricula.estado_matricula = 3  # Reprobado por calificación
            elif preinscripcion_asociada  and preinscripcion_asociada.estado_preinscripcion == 2:  # Pendiente
                matricula.estado_matricula = 9  # Aprobado - Pendiente en Formalizacion
            elif preinscripcion_asociada and preinscripcion_asociada.estado_preinscripcion == 1:  # Inscrito
                matricula.estado_matricula = 2  # Aprobado

        matricula.save()

@login_required
def actualizarEstadoMatriculasGrupos(request):

    periodo = request.session["periodo_contextualizado_id"]
    if request.method == 'GET':
        ofertas_academicas = OfertaAcademica.objects.filter(periodo_id=periodo)
        grupos = GrupoAcademico.objects.filter(horarioCurso__curso__oferta_academica__in=ofertas_academicas)
        for grupo in grupos:
            actualizarEstadoMatriculasPorGrupo(grupo, ofertas_academicas)

    return redirect(request.META['HTTP_REFERER'])

@login_required
def actualizarEstadoMatriculas(request, grupoacademico):

    periodo = request.session["periodo_contextualizado_id"]
    if request.method == 'GET':
        ofertas_academicas = OfertaAcademica.objects.filter(periodo_id=periodo)
        grupo = GrupoAcademico.objects.get(pk=grupoacademico)
        actualizarEstadoMatriculasPorGrupo(grupo, ofertas_academicas)

    return redirect(request.META['HTTP_REFERER'])


@login_required
def descargarNotasPorGrupo(request, grupoacademico):

    try:
        grupo_academico = GrupoAcademico.objects.get(pk=grupoacademico)
    except GrupoAcademico.DoesNotExist:
        grupo_academico = None

    # Matriculas: Se excluyen aquellas que estan canceladas o aplazadas por usuario y departamento
    matriculas = Matricula.objects.filter(grupo=grupo_academico).exclude(estado_matricula__in=[4,5,6]).order_by('estudiante__primer_apellido').all()
    observaciones = {}

    for matricula in matriculas:
        observaciones[matricula] = len(Observacion.objects.filter(matricula=matricula))

    data = {i+1: [str(matriculas[i].grupo.codigo) ,
                  matriculas[i].estudiante.numero_documento,
                  (matriculas[i].estudiante.getApellidos() + ' ' + matriculas[i].estudiante.getNombres()).upper(),
                  matriculas[i].estudiante.usuario.email,
                  matriculas[i].grupo.horarioCurso.curso.oferta_academica.periodo.alias,
                  matriculas[i].grupo.horarioCurso.nombre,
                  str(matriculas[i].calificacionFinal),
                  getEstadoMatricula(matriculas[i].estado_matricula),
                  str(matriculas[i].total_fallas),
                  str(observaciones[matriculas[i]])]
             for i in range(len(matriculas))}

    header = ['#','codigo-grupo', 'Documento','Nombre estudiante','Correo electronico','Periodo','Curso','Nota final','Estado', '# Inasistencias', '# Observaciones']

    csv_writer = CSVWriter()
    response = csv_writer.download_csv_file(data, header, str(grupo_academico.codigo) + '-Calificaciones')
    return response

@login_required
def descargarNotasGrupos(request):

    ofertas_academicas = OfertaAcademica.objects.filter(periodo__id=request.session['periodo_contextualizado_id'])
    grupos_academicos = GrupoAcademico.objects.filter(horarioCurso__curso__oferta_academica__in=ofertas_academicas).order_by('horarioCurso__curso__nivel__idioma__nombre', 'horarioCurso__curso__nivel__nombre')

    data = {}
    i = 1
    for grupo_academico in grupos_academicos:
        # Matriculas: Se excluyen aquellas que estan canceladas o aplazadas por usuario y departamento
        matriculas = Matricula.objects.filter(grupo=grupo_academico).exclude(estado_matricula__in=[4,5,6]).order_by('estudiante__primer_apellido').all()
        observaciones = {}

        for matricula in matriculas:
            observaciones[matricula] = len(Observacion.objects.filter(matricula=matricula))

            data[i] = [str(matricula.grupo.codigo) ,
                  matricula.estudiante.numero_documento,
                  (matricula.estudiante.getApellidos() + ' ' + matricula.estudiante.getNombres()).upper(),
                  matricula.estudiante.usuario.email,
                  matricula.grupo.horarioCurso.curso.oferta_academica.periodo.alias,
                  matricula.grupo.horarioCurso.nombre,
                  str(matricula.calificacionFinal),
                  getEstadoMatricula(matricula.estado_matricula),
                  str(matricula.total_fallas),
                  str(observaciones[matricula])]

            i += 1

    header = ['#','codigo-grupo', 'Documento','Nombre estudiante','Correo electronico','Periodo','Curso','Nota final','Estado', '# Inasistencias', '# Observaciones']

    csv_writer = CSVWriter()
    response = csv_writer.download_csv_file(data, header, 'Calificaciones')
    return response


@login_required
def generarListasExcelGruposTodos(request):
    """
    Genera archivos Excel de listas de estudiantes para todos los grupos del periodo actual,
    basándose en la estructura de plantilla_estudiantes_lista.xlsx y usando la lógica
    del elif de listadoCalificacionesPlanilla. Comprime todos los archivos Excel en un ZIP
    y genera un enlace temporal usando resultadoListasComprimido.
    """
    print("=== INICIO generarListasExcelGruposTodos ===")
    
    periodo_actual_id = request.session["periodo_contextualizado_id"]
    
    try:
        periodo = Periodo.objects.get(pk=periodo_actual_id)
        print(f"Periodo: {periodo.alias}")
    except Periodo.DoesNotExist:
        messages.error(request, 'No se encontró el periodo actual.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Obtener grupos académicos del periodo (igual que en descargarNotasGrupos)
    ofertas_academicas = OfertaAcademica.objects.filter(periodo__id=periodo_actual_id)
    grupos_academicos = GrupoAcademico.objects.filter(
        horarioCurso__curso__oferta_academica__in=ofertas_academicas
    ).order_by('horarioCurso__curso__nivel__idioma__nombre', 'horarioCurso__curso__nivel__nombre')

    if not grupos_academicos.exists():
        messages.warning(request, 'No se encontraron grupos para el periodo actual.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Configuración para archivos temporales
    excel_files = []
    unique_id = str(uuid.uuid4())[:8]
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_downloads')
    os.makedirs(temp_dir, exist_ok=True)

    grupos_con_estudiantes = 0

    try:
        # Verificar que existe la plantilla base (usa plantilla_listas.xlsx como en el elif)
        plantilla_path = os.path.join(settings.BASE_DIR, 'plantillas', 'plantilla_estudiantes_lista.xlsx')
        if not os.path.exists(plantilla_path):
            messages.error(request, 'No se encontró la plantilla base plantilla_estudiantes_lista.xlsx')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        # Procesar cada grupo (igual que en descargarNotasGrupos)
        for grupo in grupos_academicos:
            print(f"Procesando grupo: {grupo.codigo}")
            
            # Obtener matriculas (excluyendo canceladas/aplazadas)
            matriculas = Matricula.objects.filter(grupo=grupo).exclude(
                estado_matricula__in=[4, 5, 6]
            ).order_by('estudiante__primer_apellido')

            if not matriculas.exists():
                print(f"Saltando grupo {grupo.codigo} - sin estudiantes")
                continue

            grupos_con_estudiantes += 1

            # Preparar datos usando la lógica EXACTA del elif de listadoCalificacionesPlanilla
            salones = grupo.salones.all()
            docentes_grupo = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo).all()
            curso = grupo.horarioCurso.curso
            
            # Generar archivo Excel usando la MISMA lógica del elif (líneas 623-656)
            fila_inicio = 11  # Exactamente como en el elif original
            
            # Cargar plantilla Excel (igual que en el elif)
            wb = load_workbook(plantilla_path)
            ws = wb.active

            fechas = ''
            complemento_programa = ''
            salones = (f"Edificio {salones[0].edificio.nombre}  Salón: {salones[0].nombre}".upper() if len(salones) > 0 else '')

            if 'IPARM' in periodo.alias:
                fechas = 'INICIO 5 DE AGOSTO / FINALIZACIÓN 25 DE NOVIEMBRE DE 2025'
                complemento_programa = ' SEMESTRAL PRESENCIAL'
            elif 'SEM' in periodo.alias:
                fechas = 'INICIO 9 DE AGOSTO / FINALIZACIÓN 22 DE NOVIEMBRE DE 2025'
                if 'PRES'  in periodo.alias:
                    complemento_programa = ' SEMESTRAL PRESENCIAL'
                elif 'VIRT' in periodo.alias:
                    complemento_programa = ' SEMESTRAL REMOTA SINCRÓNICA'
                    salones = grupo.enlace_virtual
            elif 'BIM' in periodo.alias:
                fechas = 'INICIO 4 DE AGOSTO / FINALIZACIÓN 25 DE SEPTIEMBRE DE 2025'
                if 'PRES'  in periodo.alias:
                    complemento_programa = ' SEMESTRAL PRESENCIAL'
                elif 'VIRT' in periodo.alias:
                    salones = grupo.enlace_virtual
                    complemento_programa = ' SEMESTRAL REMOTA SINCRÓNICA'

            # Llenar datos EXACTAMENTE como en el elif de listadoCalificacionesPlanilla
            ws['B3'] = 'PROGRAMA' + ' ' + curso.oferta_academica.programa.nombre.upper() + ' ' + complemento_programa
            ws['B4'] = fechas
            ws['B5'] = 'HORARIO ' + str(humanizar_fecha.humanizar_horario(grupo.horarioCurso.horario.nombre)).upper()
            ws['B6'] = 'LISTADO DEL GRUPO - ' + str(grupo.codigo)
            ws['B7'] = 'NIVEL: ' + curso.nivel.nombre

            docente_nombre = docentes_grupo[0].docente.persona.getNombreCompleto() if len(docentes_grupo) > 0 else ''
            
            ws['A8'] = 'PROFESOR: ' + docente_nombre.upper()
            ws['A9'] = 'SALÓN: ' + salones

            # Llenar estudiantes
            contador = 0
            for idx, matricula in enumerate(matriculas, start=fila_inicio):
                contador += 1
                ws[f"A{idx}"] = contador 
                ws[f"B{idx}"] = matricula.estudiante.getApellidos().upper() + ' ' + matricula.estudiante.getNombres().upper()

            # Guardar archivo Excel individual
            excel_filename = f'{grupo.codigo}_{docente_nombre.replace(" ", "_")}_{unique_id}.xlsx'
            excel_temp_path = os.path.join(temp_dir, excel_filename)
            
            wb.save(excel_temp_path)
            wb.close()
            
            # Verificar que el archivo se generó correctamente
            if os.path.exists(excel_temp_path) and os.path.getsize(excel_temp_path) > 0:
                excel_files.append({
                    'path': excel_temp_path,
                    'name': f'{grupo.codigo}_{docente_nombre.replace(" ", "_")}.xlsx'
                })
                print(f"Excel generado exitosamente: {grupo.codigo}")
            else:
                print(f"Error: Excel vacío o no generado para grupo {grupo.codigo}")

        if not excel_files:
            messages.error(request, 'No se pudieron generar archivos Excel para ningún grupo.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Crear archivo ZIP con todos los archivos Excel
        import zipfile
        
        zip_filename = f'listas_estudiantes_todos_{periodo.alias}_{unique_id}.zip'
        zip_path = os.path.join(temp_dir, zip_filename)
        
        print(f"Creando ZIP con {len(excel_files)} archivos Excel...")
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for excel_file in excel_files:
                # Agregar archivo al ZIP con nombre limpio
                zipf.write(excel_file['path'], excel_file['name'])
                print(f"  Agregado al ZIP: {excel_file['name']}")
        
        # Limpiar archivos Excel individuales
        for excel_file in excel_files:
            try:
                os.remove(excel_file['path'])
            except:
                pass

        # Verificar que el archivo ZIP final existe
        if not os.path.exists(zip_path):
            messages.error(request, 'Error: No se pudo crear el archivo ZIP final.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        zip_size = os.path.getsize(zip_path)
        print(f"Proceso completado: {grupos_con_estudiantes} grupos procesados")
        print(f"ZIP generado: {zip_filename} ({zip_size} bytes)")
        
        # Redirigir a página de resultado usando resultadoListasComprimido
        return redirect('resultado_listas_comprimido', filename=zip_filename, grupos_count=grupos_con_estudiantes)

    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f'Error al generar los archivos Excel: {str(e)}')
        
        # Limpiar archivos temporales en caso de error
        for excel_file in excel_files:
            try:
                os.remove(excel_file['path'])
            except:
                pass
        return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def descargarArchivoTemporal(request, filename):
    """
    Descarga un archivo temporal generado previamente.
    """
    try:
        # Validar que el nombre del archivo sea seguro
        if not (filename.endswith('.pdf') or filename.endswith('.zip')) or '..' in filename or '/' in filename:
            messages.error(request, 'Nombre de archivo no válido.')
            return redirect('/')
        
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_downloads')
        temp_file_path = os.path.join(temp_dir, filename)
        
        if os.path.exists(temp_file_path):
            with open(temp_file_path, 'rb') as f:
                file_content = f.read()
            
            # Determinar el tipo de contenido basado en la extensión
            if filename.endswith('.pdf'):
                content_type = 'application/pdf'
            elif filename.endswith('.zip'):
                content_type = 'application/zip'
            else:
                content_type = 'application/octet-stream'
            
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Opcional: eliminar el archivo después de la descarga
            try:
                os.remove(temp_file_path)
            except:
                pass  # Si no se puede eliminar, continuar
            
            return response
        else:
            messages.error(request, 'El archivo solicitado no existe o ha expirado.')
            return redirect('/')
            
    except Exception as e:
        messages.error(request, f'Error al descargar el archivo: {str(e)}')
        return redirect('/')


@login_required
def resultadoListasComprimido(request, filename, grupos_count):
    """
    Muestra la página de resultado con el enlace de descarga del archivo ZIP generado.
    """
    print(f"=== RESULTADO LISTAS COMPRIMIDO ===")
    print(f"Filename: {filename}")
    print(f"Grupos count: {grupos_count}")
    
    # Validar que el archivo existe
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_downloads')
    temp_file_path = os.path.join(temp_dir, filename)
    
    print(f"Buscando archivo en: {temp_file_path}")
    print(f"Archivo existe: {os.path.exists(temp_file_path)}")
    
    if not os.path.exists(temp_file_path):
        print("ERROR: Archivo no encontrado")
        messages.error(request, f'El archivo solicitado "{filename}" no existe o ha expirado.')
        return redirect('mis-cursos')
    
    # Verificar que el archivo no esté vacío
    file_size = os.path.getsize(temp_file_path)
    print(f"Tamaño del archivo: {file_size} bytes")
    
    if file_size == 0:
        print("ERROR: Archivo vacío")
        messages.error(request, 'El archivo está vacío o corrupto.')
        return redirect('mis-cursos')
    
    # Generar URL de descarga
    download_url = reverse('descargar_archivo_temporal', kwargs={'filename': filename})
    print(f"URL de descarga: {download_url}")
    
    context = {
        'filename': filename,
        'grupos_count': grupos_count,
        'download_url': download_url,
        'file_exists': True,
        'file_size': file_size,
        'file_type': 'ZIP'  # Indicar que es un archivo ZIP
    }
    
    print(f"Context: {context}")
    print("Renderizando template...")
    
    return render(request, 'administracion/docente/resultado_listas_comprimido.html', context)


@login_required
def listadoCalificacionesPlanilla(request, format="excel", tipo='listas', *args, **kwargs):

    context = {}
    now = datetime.now()
    fecha = (now.strftime("%d/%m/%Y %H:%M:%S"))
    user = request.user
    persona = Profile.objects.filter(usuario__id=user.id).first()
    docentes = Docente.objects.filter(persona__id=persona.id)
    periodo_actual = request.session["periodo_contextualizado_id"]

    if request.method == 'GET':
        try:
            pk = kwargs.get('grupoacademico')
            grupo = get_object_or_404(GrupoAcademico, pk=pk)
        except GrupoAcademico.DoesNotExist:
            grupo = None

        try:
            periodo = Periodo.objects.get(pk=periodo_actual)
        except Periodo.DoesNotExist:
            periodo = None

        if grupo and periodo:
            salones = grupo.salones.all()
            docentes = DocentesGrupoAcademico.objects.filter(grupo_academico=grupo).all()

            curso_id = grupo.horarioCurso.curso.nivel.id
            curso = grupo.horarioCurso.curso
            niveles = Nivel.objects.filter(id=curso_id).first()

            inasistencias_matricula = {}
            matriculas = Matricula.objects.filter(grupo=grupo).exclude(estado_matricula__in=[4,5,6]).order_by('estudiante__primer_apellido')
            for matricula in matriculas:
                 inasistencias = FallaAsistencia.objects.filter(matricula=matricula).order_by('fecha').all()
                 if matricula not in inasistencias_matricula:
                    inasistencias_matricula[matricula.id] = {}
                 fallas = "cantidad_fallas"
                 inasistencias_matricula[matricula.id][fallas] = 0
                 for inasistencia in inasistencias:
                    inasistencias_matricula[matricula.id][fallas] = inasistencias_matricula[matricula.id][fallas] + int(inasistencia.cantidad_fallas)

            observaciones_matricula = {}
            for matricula in matriculas:
                 observaciones = Observacion.objects.filter(matricula=matricula).all()
                 if matricula not in observaciones_matricula:
                    observaciones_matricula[matricula.id] = list()
                 for observacion in observaciones:
                     observaciones_matricula[matricula.id].append(observacion.observacion)


            if format == "pdf":
                template_path = 'administracion/docente/mis_cursos_export.html'
                context = {'matriculas': matriculas, 'periodo': periodo, 'grupo': grupo, 'inasistencias_matricula': inasistencias_matricula,
                        'docentes': docentes, 'salones': salones, 'niveles': niveles, 'fecha':fecha, 'observaciones_matricula': observaciones_matricula}

                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'filename="report.pdf"'

                template = get_template(template_path)
                html = template.render(context)

                pisa_status = pisa.CreatePDF(
                    html, dest=response)

                if pisa_status.err:
                    return HttpResponse('Algunos errores ocurrieron <pre>' + html + '</pre>')
            elif format == "excel" and tipo =='notas':
                fila_inicio = 24
                plantilla_path = os.path.join(settings.BASE_DIR, 'plantillas', 'plantilla.xlsx')
                print(plantilla_path)
                wb = load_workbook(plantilla_path)
                ws = wb.active

                ws['L10'] = grupo.codigo
                ws['E15'] = curso.nivel.idioma.nombre + " " + str(curso.nivel.orden)
                ws['E16'] = curso.oferta_academica.programa.nombre
                ws['E17'] = str(grupo.nombre.split('-')[1])
                ws['E19'] = "05 DE AGOSTO"
                ws['V15'] = docentes[0].docente.persona.getNombreCompleto() if len(docentes) > 0 else ''
                ws['V16'] = "SEMESTRAL"
                ws['V17'] = periodo.secuencia
                ws['V19'] = "25 DE NOVIEMBRE"
                ws['AS15'] = f"Edificio: {salones[0].edificio.nombre} - Salón: {salones[0].nombre}" if len(salones) > 0 else ''
                ws['AS16'] = grupo.horarioCurso.horario.nombre if grupo.horarioCurso.horario else ''

                for idx, matricula in enumerate(matriculas, start=fila_inicio):
                    ws[f"D{idx}"] = matricula.estudiante.getApellidos().upper() + ' ' + matricula.estudiante.getNombres().upper()
                """fila_inicio = 69
                for idx, matricula in enumerate(matriculas, start=fila_inicio):
                    ws[f"D{idx}"] = matricula.estudiante.getNombreCompleto()"""
                
                # Guardar en memoria
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Crear respuesta HTTP
                response = HttpResponse(output.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=planilla'+ str(grupo.codigo)+'.xlsx'
            elif format == "excel" and tipo =='listas':
                # Lógica para generar la planilla de listas
                fila_inicio = 11
                plantilla_path = os.path.join(settings.BASE_DIR, 'plantillas', 'plantilla_listas.xlsx')

                wb = load_workbook(plantilla_path)
                ws = wb.active

                ws['B3'] = 'PROGRAMA DE IPARM INGLÉS NIÑOS SEMESTRAL PRESENCIAL IPARM'
                ws['B4'] = 'INICIO 5 DE AGOSTO / FINALIZACIÓN 25 DE NOVIEMBRE DE 2025'
                ws['B5'] = 'HORARIO: MARTES Y JUEVES 2:30 PM A 4:30 PM'
                ws['B6'] = 'LISTADO DEL GRUPO - ' + str(grupo.codigo)
                ws['B7'] = 'NIVEL: ' + curso.nivel.nombre

                docente_nombre = docentes[0].docente.persona.getNombreCompleto() if len(docentes) > 0 else ''
                
                ws['A8'] = 'PROFESOR: ' + docente_nombre.upper()
                ws['A9'] = 'SALÓN: ' + (f"Edificio {salones[0].edificio.nombre}  Salón: {salones[0].nombre}".upper() if len(salones) > 0 else '')

                contador = 0
                for idx, matricula in enumerate(matriculas, start=fila_inicio):
                    contador += 1
                    ws[f"A{idx}"] = contador 
                    ws[f"B{idx}"] = matricula.estudiante.getApellidos().upper() + ' ' + matricula.estudiante.getNombres().upper()
                    ws[f"C{idx}"] = matricula.estudiante.telefono_celular

                 # Guardar en memoria
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Crear respuesta HTTP
                response = HttpResponse(output.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename='+ str(grupo.codigo)+'_' + docente_nombre + '.xlsx'

            return response